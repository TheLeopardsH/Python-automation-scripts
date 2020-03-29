#!usr/bin/python3
#script for updating the price of Celery,Spinach,Apples in Exel Spreadsheet file
import  openpyxl
excel = openpyxl.load_workbook('produceSales.xlsx')
sheet = excel['Sheet']
#Data structure for product type and thier prices

Price_Update = {'Apples':1.88,
                'Celery':1.19,
                'Spinach':4.11}
# loop through the rows and update the prices
for Row in range(2,sheet.max_row):
    productname = sheet.cell(row=Row,column=1).value
    if productname in Price_Update:
        sheet.cell(row=Row,column=2).value = Price_Update[productname]
excel.save('UpdatedProduceSales.xlxs')
